import json
from openpyxl import Workbook
import os

workbook = Workbook()
sheet = workbook.active

sheet.cell(row=1, column=1).value = 'City'
sheet.cell(row=1, column=2).value = 'Phone number'

start_row = 2

file_names = os.listdir('result')
for file_name in file_names:
    with open('result/' + file_name, 'r') as fp:
        content = json.load(fp)
    for data in content:
        sheet.cell(row=start_row, column=1).value = data["city"]
        sheet.cell(row=start_row, column=2).value = data["phone"]
        start_row += 1
workbook.save('result.xlsx')